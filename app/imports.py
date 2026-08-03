from __future__ import annotations

import csv
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Any

from fastapi import HTTPException, status
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.constants import DEFAULT_CITY, DEFAULT_UNIT
from app.models.entities import Batch, Location, Product, Warehouse, Zone
from app.models.enums import LocationKind
from app.schemas import BatchCreate, LocationCreate, ProductCreate, WarehouseCreate, ZoneCreate
from app.services import create_batch, create_location, create_product, create_warehouse, create_zone


IMPORT_KINDS = {"products", "batches", "locations"}
HEADER_ALIASES = {
    "код": "code",
    "артикул": "code",
    "наименование": "name",
    "название": "name",
    "единица": "unit",
    "ед": "unit",
    "срок годности": "shelf_life_days",
    "товар": "product_code",
    "код товара": "product_code",
    "партия": "batch_number",
    "номер партии": "batch_number",
    "дата производства": "production_date",
    "годен до": "expiry_date",
    "срок до": "expiry_date",
    "склад": "warehouse_code",
    "код склада": "warehouse_code",
    "зона": "zone_code",
    "код зоны": "zone_code",
    "тип": "kind",
    "вместимость": "capacity_units",
}
KIND_ALIASES = {
    "хранение": LocationKind.STORAGE,
    "приемка": LocationKind.RECEIVING,
    "приёмка": LocationKind.RECEIVING,
    "карантин": LocationKind.QUARANTINE,
    "расхождения": LocationKind.DISCREPANCY,
    "экспедиция": LocationKind.EXPEDITION,
    "списание": LocationKind.SCRAP,
}


def normalize_header(value: Any) -> str:
    key = str(value or "").strip().lower().replace("\n", " ")
    key = " ".join(key.split())
    key = key.replace("-", "_")
    return HEADER_ALIASES.get(key, key)


def clean_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def parse_import_file(filename: str, content: bytes) -> list[dict[str, Any]]:
    suffix = filename.rsplit(".", 1)[-1].lower() if "." in filename else "csv"
    if suffix in {"xlsx", "xlsm"}:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    elif suffix in {"csv", "txt", "tsv"}:
        text = content.decode("utf-8-sig")
        if suffix == "tsv":
            dialect = csv.excel_tab
        else:
            try:
                dialect = csv.Sniffer().sniff(text[:2048], delimiters=",;\t")
            except csv.Error:
                dialect = csv.excel
        rows = list(csv.reader(StringIO(text), dialect))
    else:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Поддерживаются CSV, TSV и XLSX")
    if not rows:
        return []
    headers = [normalize_header(cell) for cell in rows[0]]
    parsed = []
    for raw in rows[1:]:
        row = {header: clean_cell(raw[index] if index < len(raw) else "") for index, header in enumerate(headers) if header}
        if any(value != "" for value in row.values()):
            parsed.append(row)
    return parsed


def as_int(value: Any, default: int | None = None) -> int:
    if value in ("", None):
        if default is None:
            raise ValueError("не заполнено число")
        return default
    return int(float(str(value).replace(",", ".")))


def as_date(value: Any) -> date:
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError("дата должна быть YYYY-MM-DD или DD.MM.YYYY")


def row_error(row_number: int, message: str) -> dict[str, Any]:
    return {"row_number": row_number, "message": message}


def validate_import_rows(kind: str, rows: list[dict[str, Any]], db: Session) -> dict:
    if kind not in IMPORT_KINDS:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Неизвестный тип импорта")
    errors = []
    normalized = []
    for index, row in enumerate(rows, start=2):
        try:
            if kind == "products":
                item = normalize_product_row(row)
            elif kind == "batches":
                item = normalize_batch_row(row, db)
            else:
                item = normalize_location_row(row, db)
            normalized.append({"row_number": index, "data": item})
        except ValueError as exc:
            errors.append(row_error(index, str(exc)))
    return {
        "kind": kind,
        "total_rows": len(rows),
        "valid_rows": len(normalized),
        "errors": errors,
        "rows": normalized,
    }


def normalize_product_row(row: dict[str, Any]) -> dict[str, Any]:
    code = str(row.get("code", "")).strip()
    name = str(row.get("name", "")).strip()
    if not code:
        raise ValueError("нет кода товара")
    if not name:
        raise ValueError("нет наименования товара")
    return {
        "code": code,
        "name": name,
        "unit": str(row.get("unit") or DEFAULT_UNIT),
        "shelf_life_days": as_int(row.get("shelf_life_days"), 365) if row.get("shelf_life_days") not in ("", None) else None,
    }


def normalize_batch_row(row: dict[str, Any], db: Session) -> dict[str, Any]:
    product_code = str(row.get("product_code") or row.get("product_id") or "").strip()
    batch_number = str(row.get("batch_number", "")).strip()
    if not product_code:
        raise ValueError("нет кода товара")
    if not batch_number:
        raise ValueError("нет номера партии")
    product = db.scalar(select(Product).where(Product.code == product_code))
    if product is None and product_code.isdigit():
        product = db.get(Product, int(product_code))
    if product is None:
        raise ValueError(f"товар не найден: {product_code}")
    return {
        "product_id": product.id,
        "product_code": product.code,
        "batch_number": batch_number,
        "production_date": as_date(row.get("production_date")),
        "expiry_date": as_date(row.get("expiry_date")),
        "quality_status": str(row.get("quality_status") or "released"),
        "operation_status": str(row.get("operation_status") or "allowed"),
    }


def normalize_location_kind(value: Any) -> LocationKind:
    text = str(value or LocationKind.STORAGE).strip().lower()
    if text in KIND_ALIASES:
        return KIND_ALIASES[text]
    return LocationKind(text)


def normalize_location_row(row: dict[str, Any], db: Session) -> dict[str, Any]:
    code = str(row.get("code", "")).strip()
    warehouse_code = str(row.get("warehouse_code", "")).strip()
    zone_code = str(row.get("zone_code", "")).strip()
    if not code:
        raise ValueError("нет кода ячейки")
    if not warehouse_code:
        raise ValueError("нет кода склада")
    if not zone_code:
        raise ValueError("нет кода зоны")
    kind = normalize_location_kind(row.get("kind"))
    warehouse = db.scalar(select(Warehouse).where(Warehouse.code == warehouse_code))
    zone = db.scalar(select(Zone).where(Zone.code == zone_code, Zone.warehouse_id == warehouse.id)) if warehouse else None
    return {
        "warehouse_id": warehouse.id if warehouse else None,
        "warehouse_code": warehouse_code,
        "warehouse_name": str(row.get("warehouse_name") or warehouse_code),
        "zone_id": zone.id if zone else None,
        "zone_code": zone_code,
        "zone_name": str(row.get("zone_name") or zone_code),
        "code": code,
        "name": str(row.get("name") or code),
        "kind": kind,
        "capacity_units": as_int(row.get("capacity_units"), 1),
    }


def apply_import(kind: str, rows: list[dict[str, Any]], db: Session) -> dict:
    preview = validate_import_rows(kind, rows, db)
    if preview["errors"]:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail={"message": "Есть ошибки в файле", "errors": preview["errors"]})
    created = 0
    skipped = 0
    for item in preview["rows"]:
        data = item["data"]
        if kind == "products":
            if db.scalar(select(Product.id).where(Product.code == data["code"])):
                skipped += 1
                continue
            create_product(db, ProductCreate(**data))
        elif kind == "batches":
            if db.scalar(select(Batch.id).where(Batch.product_id == data["product_id"], Batch.batch_number == data["batch_number"])):
                skipped += 1
                continue
            payload = {key: data[key] for key in ("product_id", "batch_number", "production_date", "expiry_date", "quality_status", "operation_status")}
            create_batch(db, BatchCreate(**payload))
        else:
            if db.scalar(select(Location.id).where(Location.code == data["code"])):
                skipped += 1
                continue
            warehouse = db.scalar(select(Warehouse).where(Warehouse.code == data["warehouse_code"]))
            if warehouse is None:
                warehouse = create_warehouse(db, WarehouseCreate(code=data["warehouse_code"], name=data["warehouse_name"], city=DEFAULT_CITY))
            zone = db.scalar(select(Zone).where(Zone.warehouse_id == warehouse.id, Zone.code == data["zone_code"]))
            if zone is None:
                zone = create_zone(db, ZoneCreate(warehouse_id=warehouse.id, code=data["zone_code"], name=data["zone_name"], kind=data["kind"]))
            create_location(
                db,
                LocationCreate(
                    warehouse_id=warehouse.id,
                    zone_id=zone.id,
                    code=data["code"],
                    name=data["name"],
                    kind=data["kind"],
                    capacity_units=data["capacity_units"],
                ),
            )
        created += 1
    return {**preview, "created": created, "skipped": skipped}
